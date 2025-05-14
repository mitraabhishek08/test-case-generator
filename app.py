import streamlit as st
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO

def generate_test_cases(source_table, target_table):
    # Prepare test cases: 3 rows for TC001 with same Test Case ID
    test_cases = [
        {
            "Test Case ID": "TC001",
            "Description": "Dummy description 1 for TC001",
            "Query": f"SELECT * FROM {source_table} WHERE 1=1;",
            "Expected Result": "Dummy expected result 1"
        },
        {
            "Test Case ID": "TC001",
            "Description": "Dummy description 2 for TC001",
            "Query": f"SELECT * FROM {source_table} WHERE 2=2;",
            "Expected Result": "Dummy expected result 2"
        },
        {
            "Test Case ID": "TC001",
            "Description": "Dummy description 3 for TC001",
            "Query": f"SELECT * FROM {source_table} WHERE 3=3;",
            "Expected Result": "Dummy expected result 3"
        },
        {
            "Test Case ID": "TC002",
            "Description": "Row count on target table",
            "Query": f"SELECT COUNT(*) FROM {target_table};",
            "Expected Result": "Row count should match source"
        }
    ]

    df = pd.DataFrame(test_cases)

    # Save to in-memory Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='TestCases')

    output.seek(0)

    # Post-process Excel to merge "Test Case ID" cells for TC001 rows
    wb = load_workbook(output)
    ws = wb.active

    # Merge cells in column A (Test Case ID) for rows 2 to 4 (3 rows of TC001)
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    ws.cell(row=2, column=1).alignment = Alignment(vertical='center', horizontal='center')

    # Save back to BytesIO
    output_merged = BytesIO()
    wb.save(output_merged)
    output_merged.seek(0)

    return output_merged

# Streamlit UI
st.title("Test Case Generator")

source = st.text_input("Enter source table name:")
target = st.text_input("Enter target table name:")

if st.button("Generate Test Cases"):
    if not source or not target:
        st.error("Please enter both source and target table names.")
    else:
        excel_data = generate_test_cases(source, target)
        filename = f"test_cases_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        st.success("Test cases generated successfully!")
        st.download_button(
            label="Download Excel File",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
